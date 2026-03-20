import io
from django.http import HttpResponse
from django.utils import timezone

# Intentar importar librerías de terceros, manejar error si no existen
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    Font = Alignment = PatternFill = Border = Side = None
    get_column_letter = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
except ImportError:
    colors = letter = landscape = SimpleDocTemplate = Table = TableStyle = Paragraph = Spacer = getSampleStyleSheet = ParagraphStyle = inch = None

def generar_excel_estudiantes(queryset):
    """Genera un archivo Excel con la lista de estudiantes."""
    if not openpyxl:
        return None

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Estudiantes"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezados
    headers = [
        "Matricula", "Ap. Paterno", "Ap. Materno", "Nombres", "CURP", 
        "Nivel", "Grado", "Grupo", "Estatus", "Beca (%)", "Estrato"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = thin_border

    # Datos
    for row_num, est in enumerate(queryset, 2):
        grupo_str = "S/A"
        grado_str = "S/A"
        nivel_str = "S/A"
        
        # Resolver relaciones optimizadas
        # Nota: Asumimos que queryset viene con select_related/prefetch_related
        # Pero por seguridad hacemos gets seguros
        
        inscripcion = None
        if hasattr(est, 'active_enrollment') and est.active_enrollment:
             inscripcion = est.active_enrollment[0]
        elif est.inscripciones.filter(grupo__ciclo_escolar__activo=True).exists():
             inscripcion = est.inscripciones.filter(grupo__ciclo_escolar__activo=True).first()
             
        if inscripcion and inscripcion.grupo:
            grupo_str = inscripcion.grupo.nombre
            if inscripcion.grupo.grado:
                grado_str = inscripcion.grupo.grado.nombre
                if inscripcion.grupo.grado.nivel_educativo:
                    nivel_str = inscripcion.grupo.grado.nivel_educativo.nombre
                else:
                    nivel_str = inscripcion.grupo.grado.nivel

        estrato = est.get_estrato_actual()
        estrato_nombre = estrato.nombre if estrato else "Sin Asignar"
        estado = est.get_estado_actual()
        estatus_nombre = estado.nombre if estado else "N/A"

        row = [
            est.matricula,
            est.apellido_paterno,
            est.apellido_materno,
            est.nombre,
            est.usuario.username,
            nivel_str,
            grado_str,
            grupo_str,
            estatus_nombre,
            f"{est.porcentaje_beca}%",
            estrato_nombre
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = thin_border
            if col_num in [1, 9, 10]: # Centrar matrícula y status
                cell.alignment = center_aligned

    # Ajustar ancho columnas
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    workbook.save(output)
    output.seek(0)
    return output


def generar_pdf_estudiantes(queryset):
    """Genera un archivo PDF con la lista de estudiantes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Título
    elements.append(Paragraph("Reporte General de Estudiantes", title_style))
    elements.append(Paragraph(f"Fecha de emisión: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Tabla
    data = [["Matrícula", "Nombre Completo", "Nivel/Grado/Grupo", "Estatus", "Beca"]]
    
    for est in queryset:
        grupo_full = "S/A"
        inscripcion = None
        if hasattr(est, 'active_enrollment') and est.active_enrollment:
             inscripcion = est.active_enrollment[0]
        elif est.inscripciones.filter(grupo__ciclo_escolar__activo=True).exists():
             inscripcion = est.inscripciones.filter(grupo__ciclo_escolar__activo=True).first()

        if inscripcion and inscripcion.grupo:
             g = inscripcion.grupo
             grad = g.grado.nombre if g.grado else "?"
             nivel = g.grado.nivel_educativo.nombre if (g.grado and g.grado.nivel_educativo) else "?"
             grupo_full = f"{nivel} - {grad} {g.nombre}"
        
        estado = est.get_estado_actual()
        estatus = estado.nombre if estado else "N/A"
        
        data.append([
            str(est.matricula),
            f"{est.apellido_paterno} {est.apellido_materno} {est.nombre}",
            grupo_full,
            estatus,
            f"{est.porcentaje_beca}%"
        ])
        
    table = Table(data, colWidths=[1.2*inch, 3.5*inch, 2.5*inch, 1.5*inch, 1*inch])
    
    # Estilo Tabla
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.navy),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('ALIGN', (1,1), (1,-1), 'LEFT'), # Nombre a la izquierda
    ])
    table.setStyle(style)
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generar_excel_aspirantes(queryset):
    """Genera un archivo Excel con la lista de aspirantes."""
    if not openpyxl:
        return None

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Aspirantes"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid") # Naranja para diferenciar
    center_aligned = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezados
    headers = [
        "Folio", "Ap. Paterno", "Ap. Materno", "Nombres", "Email", 
        "Nivel Interés", "Grado Interés", "Estatus Fase", "Pago"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = thin_border

    # Datos
    for row_num, asp in enumerate(queryset, 2):
        estatus_pago = "Pagado" if asp.pago_inscripcion_realizado else "Pendiente"
        
        row = [
            asp.user.folio,
            asp.apellido_paterno,
            asp.apellido_materno,
            asp.nombre,
            asp.user.email,
            asp.nivel_ingreso,
            asp.grado_ingreso,
            f"Fase {asp.proceso_finalizado_fase}",
            estatus_pago
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = thin_border
            if col_num in [1, 8]:
                cell.alignment = center_aligned

    # Ajustar ancho
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    workbook.save(output)
    output.seek(0)
    return output


def generar_pdf_reporte_financiero(data):
    """
    Genera un reporte financiero profesional en PDF.
    """
    if not SimpleDocTemplate:
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter)) # Landscape for more space
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'MainTitle', parent=styles['Title'], fontSize=20, textColor=colors.navy, spaceAfter=20
    )
    h2_style = ParagraphStyle(
        'Subtitle', parent=styles['Heading2'], fontSize=16, textColor=colors.darkblue, spaceBefore=15, spaceAfter=10
    )
    normal_style = styles['Normal']
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, textColor=colors.whitesmoke, fontName='Helvetica-Bold')

    # 1. Título y Encabezado
    periodo_label = data.get('resumen', {}).get('periodo_label', 'TOTAL')
    elements.append(Paragraph(f"ESTADO FINANCIERO E INGRESOS - PERIODO: {periodo_label}", title_style))
    elements.append(Paragraph(f"Fecha de emisión: {timezone.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
    elements.append(Spacer(1, 0.3 * inch))
    
    # 2. Resumen Ejecutivo (Tabla horizontal)
    resumen = data.get('resumen', {})
    elements.append(Paragraph("Resumen Ejecutivo de Recaudación", h2_style))
    resumen_data = [
        ["Total Recaudado", "Deuda Pendiente", "Becados (%)", "Becados (Cant.)"],
        [
            f"${resumen.get('total_recaudado', 0):,.2f}", 
            f"${resumen.get('total_deuda', 0):,.2f}", 
            f"{resumen.get('becados_pct', 0)}%", 
            str(resumen.get('becados_count', 0))
        ]
    ]
    res_table = Table(resumen_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch, 2.2*inch])
    res_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.navy),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(res_table)
    elements.append(Spacer(1, 0.5 * inch))

    # 3. Detalle de Estudiantes (Si existe y no es muy largo, o solo los primeros N)
    detalle = data.get('detalle', [])
    if detalle:
        elements.append(Paragraph("Detalle Analítico de Adeudos y Pagos", h2_style))
        
        # Reducimos columnas para el PDF (solo las más críticas)
        headers = ["Matrícula", "Nombre Estudiante", "Concepto", "Estatus", "Pagado", "Vence/Pago"]
        data_table = [headers]
        
        for d in detalle[:100]: # Limitar a 100 en PDF para que no sea infinito
            data_table.append([
                str(d['matricula']),
                f"{d['nombre']} {d['apellido_paterno']}",
                d['concepto'],
                d['estatus'],
                f"${d['monto_pagado']:,.2f}",
                d['fecha_pago']
            ])
        
        det_table = Table(data_table, colWidths=[1*inch, 2.5*inch, 2*inch, 1.2*inch, 1.2*inch, 1.1*inch])
        det_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.darkslategrey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ]))
        elements.append(det_table)
        
        if len(detalle) > 100:
            elements.append(Spacer(1, 0.1 * inch))
            elements.append(Paragraph(f"* Se muestran los primeros 100 registros de {len(detalle)}. Para el detalle completo, consulte el archivo Excel.", styles['Italic']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_excel_reporte_financiero(data):
    """
    Genera un archivo Excel profesional con un diseño elegante y contable.
    """
    if not openpyxl:
        return None

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    
    # Fuentes y Estilos
    title_font = Font(bold=True, size=18, color="1F4E78")
    subtitle_font = Font(bold=True, size=12, color="34495E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Paleta de colores profesionales (Azul marino / Gris suave)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    summary_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='BDC3C7'), 
        right=Side(style='thin', color='BDC3C7'), 
        top=Side(style='thin', color='BDC3C7'), 
        bottom=Side(style='thin', color='BDC3C7')
    )
    
    # 1. HOJA DE RESUMEN (Dashboard Ejecutivo)
    sheet_res = workbook.active
    sheet_res.title = "Resumen Ejecutivo "
    sheet_res.sheet_view.showGridLines = False

    # Logo / Título
    sheet_res.merge_cells('B2:F2')
    sheet_res['B2'].value = "ESTADO DE INGRESOS Y RECAUDACIÓN"
    sheet_res['B2'].font = title_font
    sheet_res['B2'].alignment = Alignment(horizontal="center")

    sheet_res['B3'].value = f"Periodo Reportado: {data.get('resumen', {}).get('periodo_label', 'TOTAL')}"
    sheet_res['B3'].font = subtitle_font
    sheet_res.merge_cells('B3:F3')
    sheet_res['B3'].alignment = Alignment(horizontal="center")

    # Tabla de métricas principales
    resumen = data.get('resumen', {})
    metrics = [
        ["Métrica Institucional", "Valor / Monto"],
        ["Total Recaudado (Ingresos Efectivos)", resumen.get('total_recaudado', 0)],
        ["Cartera Vencida (Deuda Pendiente)", resumen.get('total_deuda', 0)],
        ["Tasa de Estudiantes Becados", f"{resumen.get('becados_pct', 0)}%"],
        ["Población con Beca (Alumnos)", resumen.get('becados_count', 0)],
    ]

    start_row = 6
    for r_idx, row in enumerate(metrics):
        for c_idx, val in enumerate(row, 2): # Empieza en columna B
            cell = sheet_res.cell(row=start_row + r_idx, column=c_idx)
            cell.value = val
            cell.border = border_thin
            if r_idx == 0:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = summary_fill
                if c_idx == 3 and r_idx < 3: # Formato moneda para los primeros dos montos
                    cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="left" if c_idx == 2 else "right")

    # 2. HOJA DE DETALLE (EL REPORTE CONTABLE ANALÍTICO)
    sheet_det = workbook.create_sheet(title="Detalle Analítico")
    sheet_det.sheet_view.showGridLines = True
    
    # Encabezados exactos solicitados
    headers = [
        "Matrícula", "Nombre", "Apellido Paterno", "Apellido Materno", "Nivel", 
        "Adeudo (Concepto)", "Estatus del Adeudo", "Cantidad Pagada", "Fecha de Pago",
        "Tipo de Estrato", "Descuento de Estrato", "Porcentaje de Beca"
    ]
    
    # Aplicar encabezados con estilo profesional
    for c_idx, h in enumerate(headers, 1):
        cell = sheet_det.cell(row=1, column=c_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        sheet_det.row_dimensions[1].height = 25
        
    detalle = data.get('detalle', [])
    for r_idx, d in enumerate(detalle, 2):
        row_data = [
            d.get('matricula'), d.get('nombre'), d.get('apellido_paterno'), d.get('apellido_materno'),
            d.get('nivel'), d.get('concepto'), d.get('estatus'), d.get('monto_pagado'),
            d.get('fecha_pago'), d.get('tipo_estrato'), d.get('descuento_estrato'), d.get('porcentaje_beca')
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = sheet_det.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.border = border_thin
            
            # Alineación y formatos específicos
            if c_idx <= 1: # Matricula
                cell.alignment = Alignment(horizontal="center")
            if c_idx == 8: # Cantidad Pagada
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            if c_idx == 9: # Fecha
                cell.alignment = Alignment(horizontal="center")
            if c_idx >= 10: # Estratos y Becas
                cell.alignment = Alignment(horizontal="center")

    # Ajuste dinámico de columnas para legibilidad
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = 0
            if not col: continue
            column_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if cell.value:
                        l = len(str(cell.value))
                        if l > max_length: max_length = l
                except: pass
            
            # Margen extra para nombres y conceptos
            adjusted_width = max_length + 4
            if adjusted_width > 50: adjusted_width = 50 # Cap at 50 for very long text
            sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)
    return output
