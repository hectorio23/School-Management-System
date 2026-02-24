from rest_framework import status, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from academico.models import Materia
from academico.models import ProgramaEducativo

from .permissions import (
    IsAdministradorEscolar, IsMaestro, IsEstudiante
    # Nota: Los object permissions IsAdminEscolarDelMismoNivel y IsMaestroDelMismoNivel 
    # se aplicarán manualmente dentro de las funciones ya que son FBV.
)
from .models import (
    Maestro, Grupo, Materia, AsignacionMaestro, PeriodoEvaluacion,
    Calificacion, AutorizacionCambioCalificacion, ProgramaEducativo,
    CalificacionFinal, EventoCalendario, SolicitudCambioCalificacion
)
from estudiantes.models import Estudiante, CicloEscolar, Inscripcion
from .serializers import (
    MaestroSerializer, GrupoSerializer, MateriaSerializer,
    AsignacionMaestroSerializer, CalificacionSerializer,
    PeriodoEvaluacionSerializer, EstudianteSimpleSerializer,
    AsignacionWithStudentsSerializer, SolicitudCambioCalificacionSerializer
)
from .views_admin import *
from collections import OrderedDict
from decimal import Decimal
import io
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Helper para paginación
def paginate_queryset(queryset, request, serializer_class, context=None):
    paginator = PageNumberPagination()
    page = paginator.paginate_queryset(queryset, request)
    if page is not None:
        serializer = serializer_class(page, many=True, context=context)
        return paginator.get_paginated_response(serializer.data)
    serializer = serializer_class(queryset, many=True, context=context)
    return Response(serializer.data)

# =============================================================================
# VISTAS: ADMINISTRADOR ESCOLAR
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAdministradorEscolar])
def admin_maestros_list(request):
    """Listar maestros del mismo nivel educativo del administrador."""
    admin = request.user.admin_escolar_perfil
    queryset = Maestro.objects.filter(
        nivel_educativo=admin.nivel_educativo
    ).annotate(
        num_asignaciones=Count('asignaciones', filter=Q(asignaciones__activa=True))
    ).select_related('nivel_educativo', 'usuario')
    
    # Búsqueda
    search = request.query_params.get('search')
    if search:
        queryset = queryset.filter(
            Q(nombre__icontains=search) | 
            Q(apellido_paterno__icontains=search) | 
            Q(usuario__email__icontains=search)
        )
    
    serializer = MaestroSerializer(queryset, many=True)
    return Response(serializer.data)



@api_view(['GET', 'POST'])
@permission_classes([IsAdministradorEscolar])
def admin_materias_list_create(request):
    """Listar y crear materias en el nivel del administrador."""
    admin = request.user.admin_escolar_perfil
    
    if request.method == 'GET':
        queryset = Materia.objects.filter(
            grado__nivel_educativo=admin.nivel_educativo
        ).annotate(
            num_asignaciones=Count('asignaciones', filter=Q(asignaciones__activa=True))
        ).select_related('grado', 'programa_educativo')
        
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(Q(nombre__icontains=search) | Q(clave__icontains=search))
            
        serializer = MateriaSerializer(queryset, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = MateriaSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response({"status": "success", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAdministradorEscolar])
def admin_materia_desactivar(request, pk):
    """Desactivar una materia (cambio lógico)."""
    admin = request.user.admin_escolar_perfil
    materia = get_object_or_404(Materia, pk=pk)
    
    # Validación de nivel
    if materia.grado.nivel_educativo != admin.nivel_educativo:
        return Response({"status": "error", "message": "No tiene permiso sobre esta materia"}, status=403)
        
    if materia.asignaciones.filter(activa=True).exists():
         return Response(
             {"status": "error", "message": "No se puede desactivar materia con asignaciones activas"},
             status=status.HTTP_400_BAD_REQUEST
         )
    materia.activa = False
    materia.save()
    return Response({"status": "success", "message": "Materia desactivada"})

@api_view(['GET', 'POST'])
@permission_classes([IsAdministradorEscolar])
def admin_asignaciones_list_create(request):
    """Listar y crear asignaciones de maestros."""
    admin = request.user.admin_escolar_perfil
    
    if request.method == 'GET':
        queryset = AsignacionMaestro.objects.filter(
            grupo__grado__nivel_educativo=admin.nivel_educativo
        ).select_related(
            'maestro', 'grupo', 'materia', 'ciclo_escolar'
        ).annotate(
            num_calificaciones=Count('calificaciones')
        )
        serializer = AsignacionMaestroSerializer(queryset, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        data = request.data.copy()
        
        # Mapear los nombres antiguos del frontend si vienen
        if 'maestro' in data and 'maestro_id' not in data:
             data['maestro_id'] = data['maestro']
        if 'grupo' in data and 'grupo_id' not in data:
             data['grupo_id'] = data['grupo']
        if 'materia' in data and 'materia_id' not in data:
             data['materia_id'] = data['materia']
             
        # Agregar el ciclo escolar activo
        ciclo_activo = CicloEscolar.objects.filter(activo=True).first()
        if ciclo_activo:
             data['ciclo_escolar_id'] = ciclo_activo.id
             
        serializer = AsignacionMaestroSerializer(data=data)
        if serializer.is_valid():
            # Validar que los objetos pertenecen al nivel del admin
            grupo = serializer.validated_data['grupo']
            if grupo.grado.nivel_educativo != admin.nivel_educativo:
                return Response({"status": "error", "message": "El grupo no pertenece a su nivel"}, status=403)
            
            serializer.save()
            return Response({"status": "success", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAdministradorEscolar])
def admin_asignacion_detail(request, pk):
    """Obtener, modificar o eliminar una asignación específica."""
    admin = request.user.admin_escolar_perfil
    asignacion = get_object_or_404(AsignacionMaestro, pk=pk)
    
    if asignacion.grupo.grado.nivel_educativo != admin.nivel_educativo:
        return Response({"status": "error", "message": "No tiene acceso a esta asignación"}, status=403)
        
    if request.method == 'GET':
        serializer = AsignacionMaestroSerializer(asignacion)
        return Response(serializer.data)
        
    elif request.method == 'PUT':
        data = request.data.copy()
        if 'maestro' in data and 'maestro_id' not in data:
             data['maestro_id'] = data['maestro']
        if 'grupo' in data and 'grupo_id' not in data:
             data['grupo_id'] = data['grupo']
        if 'materia' in data and 'materia_id' not in data:
             data['materia_id'] = data['materia']
             
        serializer = AsignacionMaestroSerializer(asignacion, data=data, partial=True)
        if serializer.is_valid():
            if 'grupo' in serializer.validated_data:
                grupo = serializer.validated_data['grupo']
                if grupo.grado.nivel_educativo != admin.nivel_educativo:
                    return Response({"status": "error", "message": "El grupo no pertenece a su nivel"}, status=403)
            serializer.save()
            return Response({"status": "success", "data": serializer.data})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    elif request.method == 'DELETE':
        if asignacion.calificaciones.exists():
            return Response(
                 {"status": "error", "message": "No se puede eliminar porque ya tiene calificaciones capturadas."}, 
                 status=400
            )
        asignacion.delete()
        return Response({"status": "success", "message": "Asignación eliminada"}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAdministradorEscolar])
def admin_materias_disponibles(request):
    """Lista materias disponibles para un grupo (sin maestro asignado)."""
    grupo_id = request.query_params.get('grupo_id')
    if not grupo_id:
         return Response({"status": "error", "message": "grupo_id requerido"}, status=400)
    
    grupo = get_object_or_404(Grupo, pk=grupo_id)
    admin = request.user.admin_escolar_perfil
    if grupo.grado.nivel_educativo != admin.nivel_educativo:
         return Response({"status": "error", "message": "No tiene acceso a este grupo"}, status=403)

    asignadas_ids = AsignacionMaestro.objects.filter(
         grupo=grupo, activa=True
    ).values_list('materia_id', flat=True)

    materias = Materia.objects.filter(
         grado=grupo.grado, activa=True
    ).exclude(id__in=asignadas_ids)

    data = [{
         "materia_id": m.id,
         "nombre": m.nombre,
         "clave": m.clave,
         "display": f"{grupo.nombre} - {m.nombre}"
    } for m in materias]
    
    return Response({"status": "success", "data": data})

@api_view(['GET'])
@permission_classes([IsAdministradorEscolar])
def admin_calificaciones_list(request):
    """Listar todas las calificaciones registradas en el nivel con filtros."""
    admin = request.user.admin_escolar_perfil
    queryset = Calificacion.objects.filter(
        estudiante__inscripciones__grupo__grado__nivel_educativo=admin.nivel_educativo
    ).distinct().select_related('estudiante', 'asignacion_maestro__materia', 'periodo_evaluacion', 'asignacion_maestro__grupo')
    
    # Filtros
    grupo_id = request.query_params.get('grupo_id')
    matricula = request.query_params.get('matricula')
    periodo_id = request.query_params.get('periodo_id')
    
    if grupo_id:
        queryset = queryset.filter(asignacion_maestro__grupo_id=grupo_id)
    if matricula:
        queryset = queryset.filter(estudiante__matricula=matricula)
    if periodo_id:
        queryset = queryset.filter(periodo_evaluacion_id=periodo_id)
        
    return paginate_queryset(queryset, request, CalificacionSerializer)

@api_view(['POST'])
@permission_classes([IsAdministradorEscolar])
def admin_calificacion_autorizar(request, pk):
    """Autorizar cambio de una calificación bloqueada."""
    admin = request.user.admin_escolar_perfil
    calificacion = get_object_or_404(Calificacion, pk=pk)
    
    # Validar nivel
    if calificacion.estudiante.inscripciones.filter(grupo__grado__nivel_educativo=admin.nivel_educativo).exists() == False:
         return Response({"status": "error", "message": "Sin permiso sobre este estudiante"}, status=403)

    motivo = request.data.get('motivo', 'Autorizado por administrador')
    calificacion.puede_modificar = True
    calificacion.autorizada_por = admin
    calificacion.save()
    
    AutorizacionCambioCalificacion.objects.create(
        calificacion=calificacion,
        autorizado_por=admin,
        motivo=motivo,
        valor_anterior=calificacion.calificacion
    )
    
    return Response({"status": "success", "message": "Cambio autorizado para el maestro"})

@api_view(['GET', 'POST'])
@permission_classes([IsAdministradorEscolar])
def admin_periodos_list_create(request):
    """Listar y gestionar periodos de evaluación del nivel y ciclo activo."""
    admin = request.user.admin_escolar_perfil
    ciclo_activo = CicloEscolar.objects.filter(activo=True).first()
    
    if request.method == 'GET':
        queryset = PeriodoEvaluacion.objects.filter(
            programa_educativo__nivel_educativo=admin.nivel_educativo,
            ciclo_escolar=ciclo_activo
        )
        serializer = PeriodoEvaluacionSerializer(queryset, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        data = request.data.copy()
        if ciclo_activo:
            data['ciclo_escolar'] = ciclo_activo.id
            
        serializer = PeriodoEvaluacionSerializer(data=data)
        if serializer.is_valid():
            # Validar que el programa educativo pertenece al nivel del admin
            if serializer.validated_data['programa_educativo'].nivel_educativo != admin.nivel_educativo:
                 return Response({"status": "error", "message": "El programa educativo no pertenece a su nivel"}, status=403)
            
            serializer.save()
            return Response({"status": "success", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAdministradorEscolar])
def admin_periodo_detail(request, pk):
    """Obtener, modificar o eliminar un periodo específico."""
    admin = request.user.admin_escolar_perfil
    periodo = get_object_or_404(PeriodoEvaluacion, pk=pk)
    
    if periodo.programa_educativo.nivel_educativo != admin.nivel_educativo:
        return Response({"status": "error", "message": "No tiene acceso a este periodo"}, status=403)
        
    if request.method == 'GET':
        serializer = PeriodoEvaluacionSerializer(periodo)
        return Response(serializer.data)
        
    elif request.method == 'PUT':
        serializer = PeriodoEvaluacionSerializer(periodo, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": "success", "data": serializer.data})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    elif request.method == 'DELETE':
        if periodo.calificaciones.exists():
            return Response({"status": "error", "message": "No se puede eliminar porque ya tiene calificaciones capturadas."}, status=400)
        periodo.delete()
        return Response({"status": "success", "message": "Periodo eliminado"}, status=status.HTTP_204_NO_CONTENT)

# =============================================================================
# VISTAS: MAESTRO
# =============================================================================

@api_view(['GET'])
@permission_classes([IsMaestro])
def maestro_asignaciones_list(request):
    """Listar materias asignadas al maestro en el ciclo activo."""
    maestro = request.user.maestro_perfil
    queryset = AsignacionMaestro.objects.filter(
        maestro=maestro, activa=True, ciclo_escolar__activo=True
    )
    serializer = AsignacionMaestroSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsMaestro])
def maestro_estudiantes_list(request):
    """Listar estudiantes de los grupos del maestro."""
    maestro = request.user.maestro_perfil
    grupos_ids = AsignacionMaestro.objects.filter(
        maestro=maestro, activa=True, ciclo_escolar__activo=True
    ).values_list('grupo_id', flat=True)
    
    queryset = Estudiante.objects.filter(
        inscripciones__grupo_id__in=grupos_ids,
        inscripciones__estatus='activo'
    ).distinct()
    
    serializer = EstudianteSimpleSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsMaestro])
def maestro_estudiantes_grupos(request):
    """Listar asignaciones del maestro incluyendo sus estudiantes del ciclo activo."""
    maestro = request.user.maestro_perfil
    queryset = AsignacionMaestro.objects.filter(
        maestro=maestro, activa=True, ciclo_escolar__activo=True
    ).select_related('grupo', 'grupo__grado', 'materia', 'ciclo_escolar')
    
    serializer = AsignacionWithStudentsSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(['GET', 'POST'])
@permission_classes([IsMaestro])
def maestro_calificaciones_list_create(request):
    """Listar sus capturas o crear una nueva calificación."""
    maestro = request.user.maestro_perfil
    
    if request.method == 'GET':
        queryset = Calificacion.objects.filter(
            asignacion_maestro__maestro=maestro
        ).select_related('estudiante', 'asignacion_maestro__materia', 'periodo_evaluacion')
        serializer = CalificacionSerializer(queryset, many=True)
        return Response(serializer.data)
        
    elif request.method == 'POST':
        # Validar ventana de captura
        periodo_id = request.data.get('periodo_evaluacion')
        periodo = get_object_or_404(PeriodoEvaluacion, pk=periodo_id)
        today = timezone.now().date()
        
        if not (periodo.fecha_inicio_captura <= today <= periodo.fecha_fin_captura):
             return Response(
                 {"status": "error", "message": "Fuera de ventana de captura"},
                 status=status.HTTP_400_BAD_REQUEST
             )
        
        # Validar asignación
        asignacion_id = request.data.get('asignacion_maestro')
        asignacion = get_object_or_404(AsignacionMaestro, pk=asignacion_id)
        if asignacion.maestro != maestro:
             return Response({"status": "error", "message": "Asignacion incorrecta"}, status=403)

        serializer = CalificacionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(capturada_por=maestro)
            return Response({"status": "success", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsMaestro])
def maestro_solicitar_cambio(request, pk):
    """Solicitar autorización para cambiar una calificación bloqueada."""
    maestro = request.user.maestro_perfil
    calificacion = get_object_or_404(Calificacion, pk=pk)
    
    if calificacion.asignacion_maestro.maestro != maestro:
        return Response({"status": "error", "message": "No tiene permiso sobre esta calificación"}, status=403)
        
    motivo = request.data.get('motivo')
    if not motivo:
        return Response({"status": "error", "message": "El motivo es obligatorio"}, status=400)
        
    solicitud = SolicitudCambioCalificacion.objects.create(
        calificacion=calificacion,
        maestro=maestro,
        motivo=motivo
    )
    
    return Response({
        "status": "success", 
        "message": "Solicitud enviada al administrador",
        "solicitud_id": solicitud.id
    })

# --- Admin views for requests ---

@api_view(['GET'])
@permission_classes([IsAdministradorEscolar])
def admin_solicitudes_list(request):
    """Listar solicitudes de cambio de calificación pendientes para el nivel del admin."""
    admin = request.user.admin_escolar_perfil
    estatus = request.query_params.get('estatus', 'PENDIENTE')
    
    queryset = SolicitudCambioCalificacion.objects.filter(
        calificacion__estudiante__inscripciones__grupo__grado__nivel_educativo=admin.nivel_educativo
    ).distinct()
    
    if estatus:
        queryset = queryset.filter(estatus=estatus)
        
    serializer = SolicitudCambioCalificacionSerializer(queryset, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAdministradorEscolar])
def admin_solicitud_resolver(request, pk):
    """Aprobar o rechazar una solicitud de cambio."""
    admin = request.user.admin_escolar_perfil
    solicitud = get_object_or_404(SolicitudCambioCalificacion, pk=pk)
    
    # Validar acceso
    if solicitud.calificacion.estudiante.inscripciones.filter(grupo__grado__nivel_educativo=admin.nivel_educativo).exists() == False:
         return Response({"status": "error", "message": "Sin permiso sobre esta solicitud"}, status=403)
         
    accion = request.data.get('accion') # 'APROBAR' o 'RECHAZAR'
    comentario = request.data.get('comentario', '')
    
    if solicitud.estatus != 'PENDIENTE':
        return Response({"status": "error", "message": "Esta solicitud ya ha sido resuelta"}, status=400)
        
    if accion == 'APROBAR':
        solicitud.estatus = 'APROBADA'
        # Habilitar la edición de la calificación
        solicitud.calificacion.puede_modificar = True
        solicitud.calificacion.save()
    elif accion == 'RECHAZAR':
        solicitud.estatus = 'RECHAZADA'
    else:
        return Response({"status": "error", "message": "Acción no válida"}, status=400)
        
    solicitud.resuelto_por = admin
    solicitud.fecha_resolucion = timezone.now()
    solicitud.comentario_admin = comentario
    solicitud.save()
    
    return Response({"status": "success", "message": f"Solicitud {solicitud.estatus.lower()} exitosamente"})

# =============================================================================
# VISTAS: ESTUDIANTE
# =============================================================================

@api_view(['GET'])
@permission_classes([IsEstudiante])
def estudiante_historial_view(request):
    """Consultar historial académico del estudiante autenticado."""
    estudiante = request.user.perfil_estudiante
    queryset = Calificacion.objects.filter(
        estudiante=estudiante
    ).select_related('asignacion_maestro__materia', 'periodo_evaluacion', 'asignacion_maestro__maestro')
    
    serializer = CalificacionSerializer(queryset, many=True)
    return Response(serializer.data)


# =============================================================================
# VISTAS NUEVAS: CAPTURA MASIVA DE CALIFICACIONES POR GRUPO
# =============================================================================

@api_view(['GET'])
@permission_classes([IsMaestro])
def maestro_calificaciones_grupo(request, asignacion_id):
    """
    Retorna la lista de estudiantes de un grupo para captura de calificaciones.
    Los estudiantes ya calificados aparecen deshabilitados con su calificación.
    """
    maestro = request.user.maestro_perfil
    asignacion = get_object_or_404(AsignacionMaestro, pk=asignacion_id)

    if asignacion.maestro != maestro:
        return Response({"status": "error", "message": "No tiene permiso sobre esta asignación"}, status=403)

    # Resolver el grupo (en caso de inconsistencia de ciclos)
    grupo_resuelto = asignacion.get_resolved_group()

    # Obtener período de evaluación activo
    periodos = PeriodoEvaluacion.objects.filter(
        programa_educativo__nivel_educativo=grupo_resuelto.grado.nivel_educativo,
        ciclo_escolar=asignacion.ciclo_escolar
    ).order_by('numero_periodo')

    today = timezone.now().date()
    periodo_activo = None
    for periodo in periodos:
        if periodo.fecha_inicio_captura <= today <= periodo.fecha_fin_captura:
            periodo_activo = periodo
            break

    # Obtener estudiantes del grupo resuelto
    inscripciones = Inscripcion.objects.filter(
        grupo=grupo_resuelto,
        estatus='activo'
    ).select_related('estudiante', 'estudiante__usuario')

    estudiantes_data = []
    for insc in inscripciones:
        est = insc.estudiante
        calificacion_existente = None
        puede_capturar = False

        if periodo_activo:
            cal = Calificacion.objects.filter(
                estudiante=est,
                asignacion_maestro=asignacion,
                periodo_evaluacion=periodo_activo
            ).first()
            if cal:
                calificacion_existente = {
                    'id': cal.id,
                    'calificacion': float(cal.calificacion),
                    'puede_modificar': cal.puede_modificar,
                }
            else:
                puede_capturar = True

        estudiantes_data.append({
            'matricula': est.matricula,
            'nombre': est.nombre,
            'apellido_paterno': est.apellido_paterno,
            'apellido_materno': est.apellido_materno,
            'nombre_completo': f"{est.nombre} {est.apellido_paterno} {est.apellido_materno}",
            'calificacion_existente': calificacion_existente,
            'puede_capturar': puede_capturar,
        })

    return Response({
        "status": "success",
        "asignacion": {
            "id": asignacion.id,
            "grupo": asignacion.grupo.nombre,
            "materia": asignacion.materia.nombre,
            "ciclo_escolar": asignacion.ciclo_escolar.nombre,
        },
        "periodo_activo": {
            "id": periodo_activo.id,
            "nombre": periodo_activo.nombre,
            "numero_periodo": periodo_activo.numero_periodo,
            "fecha_inicio_captura": periodo_activo.fecha_inicio_captura,
            "fecha_fin_captura": periodo_activo.fecha_fin_captura,
        } if periodo_activo else None,
        "en_periodo_captura": periodo_activo is not None,
        "estudiantes": estudiantes_data,
        "total_estudiantes": len(estudiantes_data),
    })


@api_view(['POST'])
@permission_classes([IsMaestro])
def maestro_calificaciones_bulk(request):
    """
    Subida masiva de calificaciones por grupo.
    Recibe: { asignacion_id, periodo_evaluacion_id, calificaciones: [{matricula, calificacion}] }
    """
    maestro = request.user.maestro_perfil
    asignacion_id = request.data.get('asignacion_id')
    periodo_id = request.data.get('periodo_evaluacion_id')
    calificaciones_data = request.data.get('calificaciones', [])

    if not all([asignacion_id, periodo_id, calificaciones_data]):
        return Response(
            {"status": "error", "message": "Faltan campos requeridos: asignacion_id, periodo_evaluacion_id, calificaciones"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Validar asignación
    asignacion = get_object_or_404(AsignacionMaestro, pk=asignacion_id)
    if asignacion.maestro != maestro:
        return Response({"status": "error", "message": "No tiene permiso sobre esta asignación"}, status=403)

    # Validar período de captura
    periodo = get_object_or_404(PeriodoEvaluacion, pk=periodo_id)
    today = timezone.now().date()
    if not (periodo.fecha_inicio_captura <= today <= periodo.fecha_fin_captura):
        return Response(
            {"status": "error", "message": f"Fuera de ventana de captura ({periodo.fecha_inicio_captura} - {periodo.fecha_fin_captura})"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Procesar calificaciones
    creadas = 0
    errores = []

    for item in calificaciones_data:
        matricula = item.get('matricula')
        valor = item.get('calificacion')

        if not matricula or valor is None:
            errores.append({"matricula": matricula, "error": "Datos incompletos"})
            continue

        try:
            valor_decimal = Decimal(str(valor))
            if valor_decimal < 0 or valor_decimal > 10:
                errores.append({"matricula": matricula, "error": "Calificación debe estar entre 0 y 10"})
                continue

            estudiante = get_object_or_404(Estudiante, matricula=matricula)

            # Verificar que no exista calificación duplicada
            if Calificacion.objects.filter(
                estudiante=estudiante,
                asignacion_maestro=asignacion,
                periodo_evaluacion=periodo
            ).exists():
                errores.append({"matricula": matricula, "error": "Ya tiene calificación para este período"})
                continue

            Calificacion.objects.create(
                estudiante=estudiante,
                asignacion_maestro=asignacion,
                periodo_evaluacion=periodo,
                calificacion=valor_decimal,
                capturada_por=maestro
            )
            creadas += 1

        except Estudiante.DoesNotExist:
            errores.append({"matricula": matricula, "error": "Estudiante no encontrado"})
        except Exception as e:
            errores.append({"matricula": matricula, "error": str(e)})

    return Response({
        "status": "success",
        "calificaciones_creadas": creadas,
        "errores": errores,
        "total_enviadas": len(calificaciones_data),
    }, status=status.HTTP_201_CREATED if creadas > 0 else status.HTTP_400_BAD_REQUEST)


# =============================================================================
# HISTORIAL ACADÉMICO COMPLETO DEL ESTUDIANTE
# =============================================================================

@api_view(['GET'])
@permission_classes([IsEstudiante])
def estudiante_historial_completo(request):
    """
    Retorna el historial académico completo del estudiante, formateado para la tabla:
    Agrupa por ciclo escolar → grupo → materias con P1..PN, EF, CF, ES.
    """
    estudiante = request.user.perfil_estudiante

    # Identificar todos los ciclos donde el estudiante tiene inscripciones o calificaciones
    ciclos_insc = set(Inscripcion.objects.filter(estudiante=estudiante).values_list('grupo__ciclo_escolar_id', flat=True))
    ciclos_calif = set(Calificacion.objects.filter(estudiante=estudiante).values_list('asignacion_maestro__grupo__ciclo_escolar_id', flat=True))
    ciclos_final = set(CalificacionFinal.objects.filter(estudiante=estudiante).values_list('ciclo_escolar_id', flat=True))
    
    todos_ciclos_ids = ciclos_insc | ciclos_calif | ciclos_final
    
    ciclos = CicloEscolar.objects.filter(id__in=todos_ciclos_ids).order_by('-fecha_inicio')

    historial = []

    for ciclo in ciclos:
        # Intentar obtener inscripción para este ciclo
        insc = Inscripcion.objects.filter(estudiante=estudiante, grupo__ciclo_escolar=ciclo).first()
        
        if insc:
            grupo = insc.grupo
            grado = grupo.grado
            nivel = grado.nivel_educativo
            promedio_insc = float(insc.promedio_final) if insc.promedio_final else None
        else:
            # Si no hay inscripción, intentamos inferir grupo/grado de las calificaciones
            cal_ref = Calificacion.objects.filter(estudiante=estudiante, asignacion_maestro__grupo__ciclo_escolar=ciclo).first()
            if not cal_ref:
                cal_ref = CalificacionFinal.objects.filter(estudiante=estudiante, ciclo_escolar=ciclo).first()
            
            if cal_ref:
                # Si es CalificacionFinal, la materia tiene grado
                if hasattr(cal_ref, 'materia'):
                    grado = cal_ref.materia.grado
                    # El grupo es difícil de determinar sin inscripción o asignación, buscamos alguna asignación
                    asig_ref = AsignacionMaestro.objects.filter(calificaciones__estudiante=estudiante, grupo__ciclo_escolar=ciclo).first()
                    grupo = asig_ref.grupo if asig_ref else None
                else:
                    # Si es Calificacion, viene de una asignación que tiene grupo/grado
                    grupo = cal_ref.asignacion_maestro.grupo
                    grado = grupo.grado
                
                nivel = grado.nivel_educativo if grado else None
                promedio_insc = None
            else:
                continue # No debería pasar por el filtro inicial

        # Obtener programa educativo del ciclo o el activo del nivel
        # Buscamos si hay periodos definidos para este ciclo y nivel
        periodo_ref = PeriodoEvaluacion.objects.filter(ciclo_escolar=ciclo, programa_educativo__nivel_educativo=nivel).first()
        if periodo_ref:
            programa = periodo_ref.programa_educativo
        else:
            programa = ProgramaEducativo.objects.filter(nivel_educativo=nivel, activo=True).first()

        # Obtener períodos de evaluación configurados (mostrarlos aunque no existan en DB)
        num_periodos = programa.numero_periodos_evaluacion if programa else 0
        periodos_db = {p.numero_periodo: p for p in PeriodoEvaluacion.objects.filter(
            ciclo_escolar=ciclo,
            programa_educativo=programa
        )} if programa else {}

        periodos_data = []
        for i in range(1, num_periodos + 1):
            p_db = periodos_db.get(i)
            periodos_data.append({
                'numero_periodo': i,
                'nombre': p_db.nombre if p_db else f'Periodo {i}',
                'id': p_db.id if p_db else None
            })

        # Determinar materias según el plan de estudios del grado (mostramos historicas)
        materias_plan = list(Materia.objects.filter(
            grado=grado,
            programa_educativo=programa,
        ).order_by('orden'))

        # TAMBIÉN incluimos cualquier materia que tenga calificaciones registradas para este alumno/ciclo
        # (por si hubo error de captura o cambio de plan interno)
        materias_con_calif = Materia.objects.filter(
            asignaciones__calificaciones__estudiante=estudiante,
            asignaciones__grupo=grupo
        ).distinct()
        
        materias_con_final = Materia.objects.filter(
            calificaciones_finales__estudiante=estudiante,
            calificaciones_finales__ciclo_escolar=ciclo
        ).distinct()

        # Unimos y quitamos duplicados manteniendo el orden del plan
        id_materias_plan = {m.id for m in materias_plan}
        all_materias = materias_plan
        for m in materias_con_calif:
            if m.id not in id_materias_plan:
                all_materias.append(m)
                id_materias_plan.add(m.id)
        for m in materias_con_final:
            if m.id not in id_materias_plan:
                all_materias.append(m)
                id_materias_plan.add(m.id)

        materias_data = []
        numero = 1

        for materia in all_materias:
            calificaciones_periodos = {}
            
            # Intentar encontrar asignación para este grupo y materia (incluyendo inactivas)
            asig = AsignacionMaestro.objects.filter(
                grupo=grupo, 
                materia=materia, 
            ).first()

            for p_data in periodos_data:
                cal = None
                if asig:
                    cal = Calificacion.objects.filter(
                        estudiante=estudiante,
                        asignacion_maestro=asig,
                        periodo_evaluacion__numero_periodo=p_data['numero_periodo']
                    ).first()
                calificaciones_periodos[f"P{p_data['numero_periodo']}"] = float(cal.calificacion) if cal else None

            # Calificación final
            cal_final = CalificacionFinal.objects.filter(
                estudiante=estudiante,
                materia=materia,
                ciclo_escolar=ciclo
            ).first()

            # Determinar estatus
            if cal_final:
                cf = float(cal_final.calificacion_final)
                estatus_display = cal_final.estatus
            else:
                # Calcular promedio parcial si hay calificaciones
                vals = [v for v in calificaciones_periodos.values() if v is not None]
                if vals:
                    cf = round(sum(vals) / len(vals), 2)
                    estatus_display = 'CU' # En Curso
                else:
                    cf = None
                    estatus_display = 'CU' # En Curso / Sin capturar

            materias_data.append({
                'numero': numero,
                'materia': materia.nombre,
                'calificaciones': calificaciones_periodos,
                'calificacion_final': cf,
                'estatus': estatus_display,
            })
            numero += 1

        # Calcular promedio general
        promedios = [m['calificacion_final'] for m in materias_data if m['calificacion_final'] is not None]
        promedio_general = round(sum(promedios) / len(promedios), 2) if promedios else None

        historial.append({
            'ciclo_escolar': ciclo.nombre,
            'grupo': grupo.nombre,
            'grado': grado.nombre,
            'nivel_educativo': nivel.nombre,
            'programa_educativo': programa.nombre if programa else 'Sin programa',
            'periodos': [{'numero': p['numero_periodo'], 'nombre': p['nombre']} for p in periodos_data],
            'materias': materias_data,
            'promedio_general': promedio_general,
            'promedio_final_inscripcion': promedio_insc,
        })

    return Response({
        "status": "success",
        "estudiante": {
            "matricula": estudiante.matricula,
            "nombre": f"{estudiante.nombre} {estudiante.apellido_paterno} {estudiante.apellido_materno}",
        },
        "historial": historial,
    })


@api_view(['GET'])
@permission_classes([IsEstudiante])
def estudiante_calificaciones_pdf(request):
    """Genera un PDF con el historial de calificaciones del estudiante."""
    estudiante = request.user.perfil_estudiante

    # Reutilizar la lógica del historial completo
    inscripciones = Inscripcion.objects.filter(
        estudiante=estudiante
    ).select_related(
        'grupo', 'grupo__grado', 'grupo__grado__nivel_educativo', 'grupo__ciclo_escolar'
    ).order_by('-grupo__ciclo_escolar__fecha_inicio')

    # Generar PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], fontSize=16, alignment=1)
    elements.append(Paragraph(f"Historial Académico", titulo_style))
    elements.append(Spacer(1, 8))

    subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Heading2'], fontSize=12, alignment=1)
    nombre_completo = f"{estudiante.nombre} {estudiante.apellido_paterno} {estudiante.apellido_materno}"
    elements.append(Paragraph(f"{nombre_completo} - Matrícula: {estudiante.matricula}", subtitulo_style))
    elements.append(Spacer(1, 16))

    # Identificar ciclos
    ciclos_insc = set(Inscripcion.objects.filter(estudiante=estudiante).values_list('grupo__ciclo_escolar_id', flat=True))
    ciclos_calif = set(Calificacion.objects.filter(estudiante=estudiante).values_list('asignacion_maestro__grupo__ciclo_escolar_id', flat=True))
    ciclos_final = set(CalificacionFinal.objects.filter(estudiante=estudiante).values_list('ciclo_escolar_id', flat=True))
    todos_ciclos_ids = ciclos_insc | ciclos_calif | ciclos_final
    ciclos = CicloEscolar.objects.filter(id__in=todos_ciclos_ids).order_by('-fecha_inicio')

    for ciclo in ciclos:
        insc = Inscripcion.objects.filter(estudiante=estudiante, grupo__ciclo_escolar=ciclo).first()
        if insc:
            grupo = insc.grupo
            grado = grupo.grado
            nivel = grado.nivel_educativo
        else:
            cal_ref = Calificacion.objects.filter(estudiante=estudiante, asignacion_maestro__grupo__ciclo_escolar=ciclo).first()
            if not cal_ref: cal_ref = CalificacionFinal.objects.filter(estudiante=estudiante, ciclo_escolar=ciclo).first()
            if cal_ref:
                if hasattr(cal_ref, 'materia'):
                    grado = cal_ref.materia.grado
                    asig_ref = AsignacionMaestro.objects.filter(calificaciones__estudiante=estudiante, grupo__ciclo_escolar=ciclo).first()
                    grupo = asig_ref.grupo if asig_ref else None
                else:
                    grupo = cal_ref.asignacion_maestro.grupo
                    grado = grupo.grado
                nivel = grado.nivel_educativo
            else: continue

        # Obtener programa educativo del ciclo o el activo del nivel
        periodo_ref = PeriodoEvaluacion.objects.filter(ciclo_escolar=ciclo, programa_educativo__nivel_educativo=nivel).first()
        programa = periodo_ref.programa_educativo if periodo_ref else ProgramaEducativo.objects.filter(nivel_educativo=nivel, activo=True).first()

        num_periodos = programa.numero_periodos_evaluacion if programa else 0
        periodos_db = {p.numero_periodo: p for p in PeriodoEvaluacion.objects.filter(
            ciclo_escolar=ciclo,
            programa_educativo=programa
        )} if programa else {}

        periodos_data = []
        for i in range(1, num_periodos + 1):
            p_db = periodos_db.get(i)
            periodos_data.append({
                'numero_periodo': i,
                'nombre': p_db.nombre if p_db else f'P{i}',
                'id': p_db.id if p_db else None
            })


        # Materias del plan + materias con calificaciones
        materias_plan = list(Materia.objects.filter(
            grado=grado,
            programa_educativo=programa,
        ).order_by('orden'))

        materias_con_calif = Materia.objects.filter(
            asignaciones__calificaciones__estudiante=estudiante,
            asignaciones__grupo=grupo
        ).distinct()
        
        materias_con_final = Materia.objects.filter(
            calificaciones_finales__estudiante=estudiante,
            calificaciones_finales__ciclo_escolar=ciclo
        ).distinct()

        id_materias_plan = {m.id for m in materias_plan}
        all_materias = materias_plan
        for m in materias_con_calif:
            if m.id not in id_materias_plan:
                all_materias.append(m)
                id_materias_plan.add(m.id)
        for m in materias_con_final:
            if m.id not in id_materias_plan:
                all_materias.append(m)
                id_materias_plan.add(m.id)

        # Encabezado del ciclo
        seccion_style = ParagraphStyle('Seccion', parent=styles['Heading3'], fontSize=11)
        elements.append(Paragraph(
            f"Ciclo Escolar: {ciclo.nombre} | Grupo: {grupo.nombre} | Plan: {programa.nombre if programa else 'N/A'}",
            seccion_style
        ))
        elements.append(Spacer(1, 6))

        # Construir tabla
        headers = ['N', 'Materia'] + [f'P{i}' for i in range(1, num_periodos + 1)] + ['CF', 'ES']
        table_data = [headers]

        numero = 1
        for materia in all_materias:
            row = [str(numero), materia.nombre]

            # Intentar encontrar asignación para este grupo y materia (incluyendo inactivas)
            asig = AsignacionMaestro.objects.filter(
                grupo=grupo, 
                materia=materia, 
            ).first()

            for p_data in periodos_data:
                cal = None
                if asig:
                    cal = Calificacion.objects.filter(
                        estudiante=estudiante, 
                        asignacion_maestro=asig, 
                        periodo_evaluacion__numero_periodo=p_data['numero_periodo']
                    ).first()
                row.append(str(float(cal.calificacion)) if cal else '')

            cal_final = CalificacionFinal.objects.filter(
                estudiante=estudiante, materia=materia, ciclo_escolar=ciclo
            ).first()

            if cal_final:
                row.append(str(float(cal_final.calificacion_final)))
                row.append(cal_final.estatus)
            else:
                # Calcular promedio parcial si hay calificaciones
                vals = []
                # Re-check period grades for average if no final grade
                for p_data in periodos_data:
                    cal_p = None
                    if asig:
                        cal_p = Calificacion.objects.filter(
                            estudiante=estudiante, 
                            asignacion_maestro=asig, 
                            periodo_evaluacion__numero_periodo=p_data['numero_periodo']
                        ).first()
                    if cal_p:
                        vals.append(float(cal_p.calificacion))
                
                if vals:
                    avg = round(sum(vals) / len(vals), 2)
                    row.append(str(avg))
                else:
                    row.append('')
                row.append('CU')

            table_data.append(row)
            numero += 1

        # Crear tabla con estilo
        col_widths = [25, 220] + [45] * num_periodos + [45, 35]
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
        ]))
        elements.append(table)

        # Promedio
        if insc.promedio_final:
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(f"Promedio: {insc.promedio_final}", styles['Normal']))

        elements.append(Spacer(1, 16))

    # Leyenda
    elements.append(Paragraph("CF = Calificación Final | ES = Estatus | AO = Aprobado Ordinario | RP = Reprobado | CU = Cursando", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="calificaciones_{estudiante.matricula}.pdf"'
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def calendario_eventos(request):
    """
    GET /api/academico/calendario/
    Retorna los eventos del calendario filtrados por nivel educativo o globales.
    """
    if request.user.role == 'estudiante':
        try:
            estudiante = getattr(request.user, 'perfil_estudiante', None)
            if not estudiante:
                estudiante = Estudiante.objects.get(usuario=request.user)
            
            inscripcion = estudiante.inscripciones.filter(grupo__ciclo_escolar__activo=True).first()
            nivel = inscripcion.grupo.grado.nivel_educativo if inscripcion else None
            
            eventos = EventoCalendario.objects.filter(
                Q(es_global=True) | Q(nivel_educativo=nivel)
            )
        except Exception:
             eventos = EventoCalendario.objects.filter(es_global=True)
    else:
        # Para otros roles, mostrar todos (o filtrar según necesidades)
        eventos = EventoCalendario.objects.all()

    data = []
    for e in eventos:
        data.append({
            "id": e.id,
            "title": e.titulo,
            "description": e.descripcion,
            "start": e.fecha_inicio.isoformat(),
            "end": e.fecha_fin.isoformat(),
            "type": e.tipo_evento,
            "color": e.color,
            "allDay": e.fecha_inicio.time() == e.fecha_fin.time() == timezone.datetime.min.time()
        })
    
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAdministradorEscolar])
def admin_reporte_grupo_data(request):
    """
    Reporte avanzado de un grupo: promedios por materia y general.
    Query params: grupo_id (obligatorio), format (pdf, excel, json)
    """
    admin = request.user.admin_escolar_perfil
    grupo_id = request.query_params.get('grupo_id')
    export_format = request.query_params.get('format', 'json')

    if not grupo_id:
        return Response({"status": "error", "message": "grupo_id es requerido"}, status=400)

    grupo = get_object_or_404(Grupo, pk=grupo_id)
    if grupo.grado.nivel_educativo != admin.nivel_educativo:
        return Response({"status": "error", "message": "Sin permiso sobre este grupo"}, status=403)

    # Obtener materias y estudiantes
    materias = Materia.objects.filter(grado=grupo.grado, activa=True)
    inscripciones = Inscripcion.objects.filter(grupo=grupo, estatus='activo').select_related('estudiante')
    estudiantes = [i.estudiante for i in inscripciones]

    if not materias.exists() or not inscripciones.exists():
        return Response({"status": "error", "message": "No hay datos para generar el reporte"}, status=400)

    # Calcular promedios
    promedios_materias = {}

    for materia in materias:
        califs = Calificacion.objects.filter(
            asignacion_maestro__grupo=grupo,
            asignacion_maestro__materia=materia,
            estudiante__in=estudiantes
        ).values_list('calificacion', flat=True)
        
        avg = round(sum(califs) / len(califs), 2) if califs else 0
        promedios_materias[materia.nombre] = float(avg)

    promedio_general = round(sum(promedios_materias.values()) / len(promedios_materias), 2) if promedios_materias else 0

    if export_format == 'json':
        return Response({
            "grupo": grupo.nombre,
            "grado": grupo.grado.nombre,
            "promedios_por_materia": promedios_materias,
            "promedio_general": float(promedio_general),
            "total_estudiantes": len(estudiantes)
        })

    # Exportación PDF
    if export_format == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Reporte de Aprovechamiento Escolar", styles['Heading1']))
        elements.append(Paragraph(f"Grupo: {grupo.nombre} - {grupo.grado.nombre}", styles['Heading2']))
        elements.append(Spacer(1, 12))

        data = [["Materia", "Promedio"]]
        for mat, avg in promedios_materias.items():
            data.append([mat, str(avg)])
        data.append(["PROMEDIO GENERAL", str(promedio_general)])

        table = Table(data, colWidths=[300, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_grupo_{grupo.nombre}.pdf"'
        return response

    # Exportación Excel
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Grupo"

        # Encabezados
        ws.append(["Materia", "Promedio"])
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        for mat, avg in promedios_materias.items():
            ws.append([mat, avg])

        ws.append([])
        ws.append(["PROMEDIO GENERAL", promedio_general])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_grupo_{grupo.nombre}.xlsx"'
        return response

    return Response({"status": "error", "message": "Formato no soportado"}, status=400)
